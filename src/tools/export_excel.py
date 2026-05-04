# Author: T. Onkst | Date: 08132025

from __future__ import annotations

import argparse
import json
import math
import sqlite3
import sys
from pathlib import Path
from typing import Dict, Iterator, List, Optional, Tuple


EXCEL_MAX_ROWS: int = 1_048_576


def find_latest_run(runs_root: Path) -> Optional[Path]:
    if not runs_root.exists():
        return None
    candidates: List[Tuple[float, Path]] = []
    for p in runs_root.iterdir():
        if p.is_dir():
            try:
                candidates.append((p.stat().st_mtime, p))
            except Exception:
                continue
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]


def load_run_metadata(run_dir: Path) -> Dict[str, object]:
    meta_path = run_dir / "metadata.yaml"
    data: Dict[str, object] = {}
    if meta_path.exists():
        try:
            import yaml  # type: ignore
            data = yaml.safe_load(meta_path.read_text(encoding="utf-8")) or {}
        except Exception:
            data = {}
    data["run_dir"] = str(run_dir)
    return data


def list_sqlite_files(data_dir: Path) -> List[Path]:
    """Discover SQLite segment databases produced by SqliteWriter."""
    files: List[Path] = sorted(data_dir.glob("seg_*.db"))
    return files


def list_parquet_files(data_dir: Path) -> List[Path]:
    """Legacy: discover Parquet files from old-format runs."""
    files: List[Path] = []
    files.extend(sorted(data_dir.glob("data.parquet")))
    files.extend(sorted(data_dir.glob("data_*.parquet")))
    if files:
        return files
    for seg in sorted(data_dir.glob("seg_*")):
        if seg.is_dir():
            files.extend(sorted(seg.glob("*.parquet")))
    return files


def list_data_files(data_dir: Path) -> List[Path]:
    """Return SQLite files if present, otherwise fall back to Parquet."""
    files = list_sqlite_files(data_dir)
    if files:
        return files
    return list_parquet_files(data_dir)


def _is_sqlite(path: Path) -> bool:
    return path.suffix.lower() == ".db"


# ------------------------------------------------------------------
# SQLite readers
# ------------------------------------------------------------------

def get_columns_for_sqlite(path: Path) -> List[str]:
    base = ["Time_Relative_s", "Time_Absolute_iso8601"]
    others: List[str] = []
    try:
        conn = sqlite3.connect(str(path))
        try:
            for row in conn.execute("PRAGMA table_info(data)"):
                name = str(row[1])
                if name in base:
                    continue
                if name not in others:
                    others.append(name)
        finally:
            conn.close()
    except Exception:
        pass
    return list(base) + sorted(others)


def iter_sqlite_dataframes(
    path: Path, columns: List[str], chunk_size: int = 50_000
) -> Iterator:
    """Yield pandas DataFrames in chunks from a SQLite segment database."""
    try:
        import pandas as pd  # type: ignore
    except Exception as e:
        raise RuntimeError(f"pandas required to export SQLite: {e}")
    col_expr = ", ".join(f'"{c}"' for c in columns)
    conn = sqlite3.connect(str(path))
    try:
        for chunk_df in pd.read_sql(
            f"SELECT {col_expr} FROM data", conn, chunksize=chunk_size
        ):
            yield chunk_df
    finally:
        conn.close()


def read_units_metadata_sqlite(path: Path) -> Dict[str, str]:
    try:
        conn = sqlite3.connect(str(path))
        try:
            cur = conn.execute(
                "SELECT value FROM _metadata WHERE key = ?", ("units_json",)
            )
            row = cur.fetchone()
            if row:
                return json.loads(row[0])
        finally:
            conn.close()
    except Exception:
        pass
    return {}


# ------------------------------------------------------------------
# Parquet readers (kept for backward compatibility with old runs)
# ------------------------------------------------------------------

def get_columns_for_parquet(path: Path) -> List[str]:
    base = ["Time_Relative_s", "Time_Absolute_iso8601"]
    others: List[str] = []
    try:
        import pyarrow.parquet as pq  # type: ignore
        pf = pq.ParquetFile(path)
        for name in pf.schema.names:
            if name in base:
                continue
            if name not in others:
                others.append(name)
    except Exception:
        try:
            import pandas as pd  # type: ignore
            cols = list(pd.read_parquet(path).columns)
            for c in cols:
                if c in base:
                    continue
                if c not in others:
                    others.append(c)
        except Exception:
            pass
    return list(base) + sorted(others)


def iter_parquet_dataframes(path: Path, columns: List[str]):
    """Yield pandas DataFrames per row group from a Parquet file."""
    try:
        import pyarrow as pa  # type: ignore
        import pyarrow.parquet as pq  # type: ignore
        import pandas as pd  # type: ignore
    except Exception as e:
        raise RuntimeError(f"pyarrow/pandas required to export Parquet: {e}")
    pf = pq.ParquetFile(path)
    for rg_index in range(pf.num_row_groups):
        try:
            tbl = pf.read_row_group(rg_index)
        except Exception:
            tbl = pf.read()
        keep = [c for c in columns if c in set(tbl.column_names)]
        drop = [c for c in tbl.column_names if c not in keep]
        if drop:
            tbl = tbl.drop(drop)
        missing = [c for c in columns if c not in set(tbl.column_names)]
        if missing:
            arrays = []
            for m in missing:
                arrays.append(pa.nulls(len(tbl)))
            add = pa.table({missing[i]: arrays[i] for i in range(len(missing))})
            tbl = pa.concat_tables([tbl, add], promote=True)
        tbl = tbl.select(columns)
        df = tbl.to_pandas(types_mapper=None)
        yield df


# ------------------------------------------------------------------
# Unified dispatch helpers
# ------------------------------------------------------------------

def get_columns_for_file(path: Path) -> List[str]:
    if _is_sqlite(path):
        return get_columns_for_sqlite(path)
    return get_columns_for_parquet(path)


def iter_dataframes(path: Path, columns: List[str]):
    if _is_sqlite(path):
        yield from iter_sqlite_dataframes(path, columns)
    else:
        yield from iter_parquet_dataframes(path, columns)


def write_metadata_sheet(writer, run_meta: Dict[str, object], data_files: List[Path], total_rows: int, units_sample: Dict[str, str]) -> None:
    try:
        import pandas as pd  # type: ignore
    except Exception:
        return
    rows: List[Tuple[str, object]] = []
    rows.append(("run_dir", run_meta.get("run_dir", "")))
    rows.append(("run_id", run_meta.get("run_id", Path(str(run_meta.get("run_dir", ""))).name)))
    rows.append(("recording_rate_hz", run_meta.get("recording_rate_hz", "")))
    rows.append(("plugins", ", ".join(run_meta.get("plugins", [])) if isinstance(run_meta.get("plugins"), list) else run_meta.get("plugins", "")))
    rows.append(("data_files", ", ".join([f.name for f in data_files])))
    rows.append(("total_rows", total_rows))
    df = pd.DataFrame(rows, columns=["key", "value"])
    df.to_excel(writer, sheet_name="Metadata", index=False)
    # Write units matrix below metadata as two columns: channel, unit
    if units_sample:
        items = list(units_sample.items())
        df_units = pd.DataFrame(items, columns=["channel", "unit"])
        # place with a gap of 2 rows after the first table
        startrow = len(rows) + 2
        df_units.to_excel(writer, sheet_name="Metadata", index=False, startrow=startrow)


def _autosize_and_format_numeric(writer, sheet_name: str, df_columns: List[str], engine: str) -> None:
    try:
        import numpy as _np  # type: ignore
    except Exception:
        _np = None
    try:
        ws = writer.sheets.get(sheet_name)
    except Exception:
        ws = None
    if ws is None:
        return
    # Determine approximate widths from header
    col_widths = {i: len(str(col)) + 2 for i, col in enumerate(df_columns)}
    # Best-effort expansion: we cannot easily read back all cells without extra cost; headers-only sizing is acceptable
    # Apply widths and numeric formats
    try:
        if engine == "openpyxl":
            from openpyxl.utils import get_column_letter  # type: ignore
            from openpyxl.styles import numbers  # type: ignore
            max_row = getattr(ws, 'max_row', 0)
            for i, col in enumerate(df_columns, start=1):
                letter = get_column_letter(i)
                try:
                    ws.column_dimensions[letter].width = max(col_widths.get(i-1, 10), 10)
                except Exception:
                    pass
                # Apply 2-decimal display to numeric columns (skip first two time columns)
                if i > 2:
                    try:
                        for r in range(2, max_row + 1):
                            cell = ws.cell(row=r, column=i)
                            # Only set if value is number
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00'
                    except Exception:
                        pass
        else:
            # xlsxwriter
            try:
                wb = writer.book
                numfmt = wb.add_format({'num_format': '0.00'})
            except Exception:
                numfmt = None
            for i, col in enumerate(df_columns):
                width = max(col_widths.get(i, 10), 10)
                fmt = numfmt if (i > 1 and numfmt is not None) else None
                try:
                    ws.set_column(i, i, width, fmt)
                except Exception:
                    pass
    except Exception:
        pass


def load_units_merged(files: List[Path]) -> Dict[str, str]:
    merged: Dict[str, str] = {}
    for f in files:
        try:
            m = read_units_metadata(f)
            if isinstance(m, dict):
                for k, v in m.items():
                    if k not in merged:
                        merged[k] = v
        except Exception:
            continue
    return merged


def read_units_metadata(path: Path) -> Dict[str, str]:
    if _is_sqlite(path):
        return read_units_metadata_sqlite(path)
    return _read_units_metadata_parquet(path)


def _read_units_metadata_parquet(parquet_path: Path) -> Dict[str, str]:
    try:
        import pyarrow.parquet as pq  # type: ignore
        table = pq.read_table(parquet_path)
        md = table.schema.metadata or {}
        raw = md.get(b"units_json")
        if not raw:
            return {}
        try:
            return json.loads(raw.decode("utf-8"))
        except Exception:
            return {}
    except Exception:
        return {}


def read_jsonl(path: Path) -> Optional[object]:
    try:
        import pandas as pd  # type: ignore
        if not path.exists():
            return None
        return pd.read_json(path, lines=True)
    except Exception:
        return None


def export_excel(
    run_dir: Path,
    engine: str = "openpyxl",
    rows_per_file: int = EXCEL_MAX_ROWS - 1,
    output_dir: Optional[Path] = None,
) -> List[Path]:
    try:
        import pandas as pd  # type: ignore
    except Exception as e:
        raise RuntimeError(f"pandas required to export Excel: {e}")

    data_dir = run_dir / "data"
    if not data_dir.exists():
        raise FileNotFoundError(f"Data folder missing: {data_dir}")

    files = list_data_files(data_dir)
    if not files:
        raise FileNotFoundError(f"No data files found under {data_dir}")

    run_meta = load_run_metadata(run_dir)
    target_dir = Path(output_dir).resolve() if output_dir is not None else (run_dir / "data").resolve()
    target_dir.mkdir(parents=True, exist_ok=True)

    created: List[Path] = []

    multi_seg = len(files) > 1
    all_units = load_units_merged(files)

    for file_idx, f in enumerate(files, start=1):
        cols = get_columns_for_file(f)
        units = read_units_metadata(f) or all_units
        total_rows = 0
        part_idx = 1
        remaining = max(1, int(rows_per_file))
        frames: List["pd.DataFrame"] = []

        def flush_one(part: int, frames_list: List["pd.DataFrame"]) -> None:
            out_stem = f"Data_{run_dir.name}"
            if multi_seg:
                out_stem = f"{out_stem}_seg{file_idx}"
            out_name = f"{out_stem}.xlsx" if part == 1 and total_rows <= rows_per_file else f"{out_stem}.{part}.xlsx"
            out_path = target_dir / out_name
            with pd.ExcelWriter(out_path, engine=engine) as writer:
                if frames_list:
                    df = pd.concat(frames_list, ignore_index=True)
                    df.to_excel(writer, sheet_name="Data", index=False)
                    try:
                        _autosize_and_format_numeric(writer, "Data", list(df.columns), engine)
                    except Exception:
                        pass
                write_metadata_sheet(writer, run_meta, [f], total_rows, units)
                ae_df = read_jsonl(run_dir / "alarm_events.jsonl")
                if ae_df is not None:
                    try:
                        ae_df.to_excel(writer, sheet_name="AlarmEvents", index=False)
                    except Exception:
                        pass
            created.append(out_path)

        for df in iter_dataframes(f, cols):
            if df.empty:
                continue
            df_len = int(len(df))
            total_rows += df_len
            start = 0
            while start < df_len:
                can_take = min(remaining, df_len - start)
                frames.append(df.iloc[start:start + can_take])
                remaining -= can_take
                start += can_take
                if remaining <= 0:
                    flush_one(part_idx, frames)
                    part_idx += 1
                    frames = []
                    remaining = max(1, int(rows_per_file))

        if frames:
            flush_one(part_idx, frames)

    # Export stats snapshots as a separate file, one tab per metric
    ss_jsonl = run_dir / "stats_snapshots.jsonl"
    ss_df = read_jsonl(ss_jsonl)
    if ss_df is not None and not ss_df.empty:
        ss_path = target_dir / f"Statistics_{run_dir.name}.xlsx"
        _METRIC_SUFFIXES = [
            ("mean", "Mean"),
            ("stdev", "Std Dev"),
            ("min", "Min"),
            ("max", "Max"),
            ("p2p", "Peak-to-Peak"),
        ]
        try:
            with pd.ExcelWriter(ss_path, engine=engine) as sw:
                for suffix, sheet_name in _METRIC_SUFFIXES:
                    metric_cols = [c for c in ss_df.columns if c.endswith(f"_{suffix}")]
                    if not metric_cols:
                        continue
                    time_col = ["ts_hms"] if "ts_hms" in ss_df.columns else []
                    rename_map = {c: c.rsplit(f"_{suffix}", 1)[0] for c in metric_cols}
                    tab_df = ss_df[time_col + metric_cols].copy()
                    tab_df.rename(columns=rename_map, inplace=True)
                    tab_df.to_excel(sw, sheet_name=sheet_name, index=False)
                    try:
                        _autosize_and_format_numeric(sw, sheet_name, list(tab_df.columns), engine)
                    except Exception:
                        pass
            created.append(ss_path)
        except Exception:
            pass

    return created


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Export run data to Excel (Metadata, Data, AlarmEvents, StatsSnapshots)")
    parser.add_argument("--run", type=str, default=None, help="Path to a specific run folder (e.g., runs/081325_140121)")
    parser.add_argument("--engine", type=str, default="openpyxl", choices=["openpyxl", "xlsxwriter"], help="Excel writer engine")
    parser.add_argument("--rows-per-file", type=int, default=EXCEL_MAX_ROWS - 1, help="Max data rows per workbook before splitting")
    parser.add_argument("--output-dir", type=str, default=None, help="Destination folder for workbooks (default: <run>/data)")
    args = parser.parse_args(argv)

    project_root = Path(__file__).resolve().parents[2]
    runs_root = project_root / "runs"
    run_dir = Path(args.run).resolve() if args.run else find_latest_run(runs_root)
    if run_dir is None or not run_dir.exists():
        print("[ERROR] Run folder not found.")
        return 2

    out_dir = Path(args.output_dir).resolve() if args.output_dir else None
    try:
        outputs = export_excel(run_dir, engine=args.engine, rows_per_file=args.rows_per_file, output_dir=out_dir)
    except Exception as e:
        print(f"[ERROR] Export failed: {e}")
        return 1

    print("Exported:")
    for p in outputs:
        print(f"  - {p}")
    return 0


if __name__ == "__main__":
    sys.exit(main())


